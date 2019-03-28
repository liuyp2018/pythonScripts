from openpyxl import load_workbook,Workbook
from xlrd import open_workbook
# ************************************* #
# ***利用openpyxl对excel进行读写操作*** #
# ************************************* #
'''
# ****读excel文件操作****#
# 打开文件
wb=load_workbook('E:/test.xlsx')
print(wb.sheetnames)
# 选择sheet
sheet=wb['Sheet3']
# 获取指定单元格的值
print(sheet["A2"].value)
# 循环获取指定列的所有值
for i in sheet['B']:
    print(i.value,end=' ')
# 循环获取指定行的所有值
for j in sheet["2"]:
    print(j.value,end=' ')

# ******往excel中写入文件******#
# 创建一个工作表
wb=Workbook()
# 获取活动sheet
sheet=wb.active
# 重命名sheet
sheet.title='new sheet'
# 指定的单元格中写入数据
sheet['C3']='hello world'
sheet['A1'].value='aaa'
#保存工作表
wb.save('自动创建1.xlsx')
'''

# ************************************* #
# ***利用xlrd对excel进行读写操作*** #
# ************************************* #
#打开一个excel文件
# wd=open_workbook('E:/test.xlsx')
# #依据sheet索引，打开指定的工作表
# #sheet=wd.sheets()[0]
# #依据sheet名称，打开指定的工资表
# sheet=wd.sheet_by_name('Sheet3')
# #获取单元格数据
# print(sheet.cell(0, 0).value)
# #获取整行数据
# print(sheet.row_values(0))
# #获取整列数据
# print(sheet.col_values(0))
#
# print(sheet.cell_value(0,0))

# #获取指定用例的数据
# def get_test_data(sheetname,testcasename):
#     testdata=[]
#     sheet = wd.sheet_by_name(sheetname)
#     sheet_row_number=sheet.nrows
#     for i in range(sheet_row_number):
#         cell_value=sheet.cell_value(i,0)
#         if(cell_value==testcasename):
#             #print(sheet.row_values(i))
#             testdata.append(sheet.row_values(i))
#     return testdata
#
# data=get_test_data('Sheet3','a1')
# print(data)
# for d in data:
#     print(d)


import requests
import re
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
import shutil
import pymysql
URL='http://192.168.10.42:90'
CASE_NUMBER=0
CASE_NAME=1
CASE_URL=2
CASE_METHOD=3
CASE_AUTH=4
CASE_HEADER=5
CASE_SQL=6
CASE_DATA=7
CASE_ACTIVE=8
CASE_CHECKPOINT=9
CASE_RESPONSE=10
CASE_RESULT=11

testuser='apiTestStu1@111.com'
testteacher='apiTestTeacher1@111.com'
manager='apiTestManager1@111.com'
class_id='1593'
#数据库配置
config = {
            'host': '192.168.10.42',
            'user': 'root',
            'passwd': 'simple@213',
            'db': 'simpleove_base',
            'port': 3306,
            'charset': 'utf8'
        }


# 登陆函数
def login(username, passwd):
    data = {'username': username, 'password': passwd}
    r = requests.post(URL + '/api/login', data=data)
    return r.json()

# 获取登录用户的token-带bearer
def getToken(username, passwd='123456'):
    r = login(username, passwd)
    token = r['data']['token']
    return 'Bearer ' + token

stu_token=getToken(testuser,'123456')
t_token=getToken(testteacher,'123456')
m_token=getToken(manager,'123456')
domain='xp.ove.local:6000'
def get_cell_content(sheet,row,col):
    """
    获取表格中内容
    :param sheet: 工作表
    :param row: 行
    :param col: 列
    """
    return sheet.cell_value(row,col)

testdata_file='D:/pythonProject/interfaceTest/testfiles/testdatas.xlsx'
resultdata_file='D:/pythonProject/interfaceTest/testfiles/resultdatas.xlsx'

#利用shutil复制excel
shutil.copyfile(testdata_file,resultdata_file)

def write_excel(file,row,col,value):
    wb=openpyxl.load_workbook(file)
    ws=wb.active
    ws.cell(row,col).value=value
    wb.save(file)


#数据库操作
#连接数据库
def database():
    try:
        conn=pymysql.connect(**config)
        return conn
    except pymysql.Error as e:
        print('数据库连接失败',str(e))

def getdata_by_sql(conn,sql):
    conn.commit()
    cur=conn.cursor()
    cur.execute(sql)
    try:
        rs=cur.fetchone()
        return rs[0]
    except Exception as e:
        return None

try:
    wd=open_workbook(testdata_file)
except Exception as e:
    print('无法打开测试数据文件，请检查文件是否正常')
try:
    # 依据sheet名称，打开指定的工作表
    sheet=wd.sheet_by_name('Sheet1')
except Exception as e:
    print('打开工作表失败')
conn=database()
for i in range(1,sheet.nrows):
    # 获取用例序号
    number = int(get_cell_content(sheet, i, CASE_NUMBER))
    # 获取用例名称
    casename = get_cell_content(sheet, i, CASE_NAME)
    print('执行用例：%s-%s' % (number, casename))
    try:
        caseActive = get_cell_content(sheet, i, CASE_ACTIVE)
        if caseActive=='N':
            continue
        caseURL=get_cell_content(sheet,i,CASE_URL)
        caseMethod=get_cell_content(sheet,i,CASE_METHOD)
        caseAuth=get_cell_content(sheet,i,CASE_AUTH)
        if caseAuth=='Y':
            caseHeader=eval(get_cell_content(sheet,i,CASE_HEADER))
            print(caseHeader)
        casesql = get_cell_content(sheet, i, CASE_SQL)
        print(casesql)
        if casesql != '':
            sql_data=getdata_by_sql(conn, casesql)
            print(sql_data)
        try:
            caseData=eval(get_cell_content(sheet,i,CASE_DATA))
            print(caseData)
        except Exception as e:
            print(e)

        caseCheckpoint=get_cell_content(sheet,i,CASE_CHECKPOINT)
        url = URL + caseURL

        try:
            if caseAuth=='Y':
                if caseMethod == 'post':
                    r = requests.post(url, data=caseData, headers=caseHeader)
                elif caseMethod == 'get':
                    r = requests.get(url, params=caseData,headers=caseHeader)
            elif caseAuth=='N':
                if caseMethod == 'post':
                    r = requests.post(url, data=caseData)
                elif caseMethod == 'get':
                    r = requests.get(url, params=caseData)
            print(r.json())
            write_excel(resultdata_file, i + 1, CASE_RESPONSE + 1, str(r.json()))
            c = re.search(caseCheckpoint, str(r.json()))
            if c is not None:
                print(c)
                print('测试验证通过')
                write_excel(resultdata_file, i + 1, CASE_RESULT + 1, '验证通过')
            else:
                print('测试验证失败')
                write_excel(resultdata_file, i + 1, CASE_RESULT + 1, '验证失败')
        except Exception as e:
            print('接口请求失败')
            write_excel(resultdata_file, i + 1, CASE_RESULT + 1, '接口请求失败，未测试验证')
    except Exception as e:
        print('读取测试数据失败')
        write_excel(resultdata_file, i + 1, CASE_RESULT + 1, '读取测试数据失败，未测试验证')





# #复制excel表格方法
# def copy_excel(oldfile,newfile):
#     #创建一个excel表格
#     wb2=openpyxl.Workbook()
#     wb2.save(newfile)
#
#     wb1=openpyxl.load_workbook(oldfile)
#     wb2=openpyxl.load_workbook(newfile)
#     sheets1=wb1.sheetnames
#     sheets2=wb2.sheetnames
#     sheet1=wb1[sheets1[0]]
#     sheet2=wb2[sheets2[0]]
#     max_row=sheet1.max_row
#     max_col=sheet1.max_column
#
#     for m in range(1,max_row+1):
#         for n in range(97,97+max_col):
#             n=chr(n)
#             i='%s%s'%(n,m)
#             cell_value=sheet1[i].value
#             sheet2[i]=cell_value
#     wb2.save(newfile)
#     wb1.close()
#     wb2.close()

#copy_excel(testdata_file,resultdata_file)

#write_excel(resultdata_file,2,9)